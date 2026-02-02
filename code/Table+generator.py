import pandas as pd
import numpy as np
from pathlib import Path


IN_PATH = Path(r"Q&A.xlsx")  
OUT_PATH = Path(r"replication_table2_table3_results.xlsx")


def to_binary_series(s: pd.Series) -> pd.Series:
    """Coerce to 0/1 with NaNs preserved. Accepts strings like '1','0','yes','no'."""
    x = s.copy()
    if x.dtype == "object":
        x = (
            x.astype(str)
            .str.strip()
            .replace({"": np.nan, "nan": np.nan, "None": np.nan, "NA": np.nan, "NaN": np.nan})
        )
        x = x.replace(
            {
                "Yes": 1, "yes": 1, "Y": 1, "True": 1, "true": 1,
                "No": 0, "no": 0, "N": 0, "False": 0, "false": 0,
            }
        )
        x = pd.to_numeric(x, errors="coerce")
    return x


def confusion_metrics(y_true: pd.Series, y_pred: pd.Series) -> dict:
    """Compute confusion matrix + paper-style metrics. Positive class = 1 (non-answer)."""
    mask = y_true.notna() & y_pred.notna()
    yt = y_true[mask].astype(int)
    yp = y_pred[mask].astype(int)

    TP = int(((yt == 1) & (yp == 1)).sum())
    FP = int(((yt == 0) & (yp == 1)).sum())
    TN = int(((yt == 0) & (yp == 0)).sum())
    FN = int(((yt == 1) & (yp == 0)).sum())
    N = int(len(yt))

    acc = (TP + TN) / N if N else np.nan
    type1 = FP / (FP + TN) if (FP + TN) else np.nan  # FP rate
    type2 = FN / (FN + TP) if (FN + TP) else np.nan  # FN rate
    prec = TP / (TP + FP) if (TP + FP) else np.nan
    rec = TP / (TP + FN) if (TP + FN) else np.nan

    # Paper's F1: TP / (TP + 0.5*(FP+FN))  (equals standard F1 when defined)
    f1 = TP / (TP + 0.5 * (FP + FN)) if (TP + 0.5 * (FP + FN)) else np.nan

    # For single-label binary classification, micro avg precision/recall/F1 == accuracy
    return {
        "TP": TP, "FP": FP, "TN": TN, "FN": FN, "N": N,
        "Accuracy": acc,
        "Type I error": type1,
        "Type II error": type2,
        "Non-answers: Precision": prec,
        "Non-answers: Recall": rec,
        "Non-answers: F1 score": f1,
        "Total: Precision": acc,
        "Total: Recall": acc,
        "Total: F1 score": acc,
    }


def summarize_binary(series: pd.Series) -> dict:
    """Counts of 0/1 ignoring NaN."""
    s = series.dropna().astype(int)
    return {"Answer": int((s == 0).sum()), "Non-answer": int((s == 1).sum()), "N": int(len(s))}


def desc_stats(x: pd.Series) -> dict:
    """Descriptive stats matching Table 2 format."""
    x = x.dropna().astype(float)
    if len(x) == 0:
        return {
            "Obs": 0, "Mean": np.nan, "Std_Dev": np.nan,
            "P5": np.nan, "P25": np.nan, "P50": np.nan, "P75": np.nan, "P95": np.nan,
        }
    return {
        "Obs": int(len(x)),
        "Mean": float(x.mean()),
        "Std_Dev": float(x.std(ddof=1)) if len(x) > 1 else 0.0,
        "P5": float(np.percentile(x, 5)),
        "P25": float(np.percentile(x, 25)),
        "P50": float(np.percentile(x, 50)),
        "P75": float(np.percentile(x, 75)),
        "P95": float(np.percentile(x, 95)),
    }


def detect_prediction_columns(df: pd.DataFrame, manual_col: str) -> list:
    """Detect binary 0/1 columns excluding obvious id/text columns."""
    core_like = {manual_col}
    exclude_names = {
        "transcriptid", "qid", "question", "answer",
        "transcriptid-qid", "transcriptid_qid"
    }
    for c in df.columns:
        if str(c).strip().lower() in exclude_names:
            core_like.add(c)

    pred_cols = []
    for c in df.columns:
        if c in core_like:
            continue
        s = to_binary_series(df[c])
        vals = s.dropna().unique()
        if len(vals) and set(vals).issubset({0, 1}):
            pred_cols.append(c)

    return [c for c in df.columns if c in pred_cols]


df = pd.read_excel(IN_PATH, engine="openpyxl")
df.columns = [str(c).strip() for c in df.columns]

manual_col = next((c for c in df.columns if c.lower() == "manual"), None)
if manual_col is None:
    raise ValueError("Cannot find 'Manual' column (case-insensitive).")

pred_cols = detect_prediction_columns(df, manual_col)
if len(pred_cols) == 0:
    raise ValueError("No binary prediction columns detected (0/1).")



# Table 2 (was Table 1): Manual non-missing

manual = to_binary_series(df[manual_col])
df_eval = df.loc[manual.notna()].copy()

y_true = to_binary_series(df_eval[manual_col])

table2_rows = [
    "Answer", "Non-answer",
    "Accuracy", "Type I error", "Type II error",
    "Non-answers: Precision", "Non-answers: Recall", "Non-answers: F1 score",
    "Total: Precision", "Total: Recall", "Total: F1 score",
    "N",
]

table2 = pd.DataFrame(index=table2_rows)

# Manual counts
manual_counts = summarize_binary(y_true)
table2["Manual"] = [
    manual_counts.get(r, np.nan) if r in ["Answer", "Non-answer", "N"] else np.nan
    for r in table2_rows
]
table2.loc["N", "Manual"] = manual_counts["N"]

confusion_rows = []
for c in pred_cols:
    y_pred = to_binary_series(df_eval[c])
    cnts = summarize_binary(y_pred)
    m = confusion_metrics(y_true, y_pred)

    col_vals = {}
    col_vals["Answer"] = cnts["Answer"]
    col_vals["Non-answer"] = cnts["Non-answer"]
    for k in [
        "Accuracy", "Type I error", "Type II error",
        "Non-answers: Precision", "Non-answers: Recall", "Non-answers: F1 score",
        "Total: Precision", "Total: Recall", "Total: F1 score",
    ]:
        col_vals[k] = m[k]
    col_vals["N"] = m["N"]

    table2[c] = [col_vals.get(r, np.nan) for r in table2_rows]
    confusion_rows.append({"Method": c, **{k: m[k] for k in ["TP", "FP", "TN", "FN", "N"]}})

confusion_df = pd.DataFrame(confusion_rows).set_index("Method")

# Format Table 2 for Excel display (avoid dtype crash)
table2_fmt = table2.copy().astype("object")  

for r in table2_rows:
    if r in ["Answer", "Non-answer", "N"]:
        table2_fmt.loc[r] = table2_fmt.loc[r].apply(lambda v: "" if pd.isna(v) else int(v))
    else:
        table2_fmt.loc[r] = table2_fmt.loc[r].apply(lambda v: "" if pd.isna(v) else round(float(v), 2))



# Table 3 (was Table 2): Full sample, pair level only

pair_stats = {}
for c in pred_cols:
    pair_stats[f"{c} - % non-answer"] = desc_stats(to_binary_series(df[c]))
table3_pair = pd.DataFrame(pair_stats).T



# Write Excel (Table2_eval, Confusion_eval, Table3_pair_level)

with pd.ExcelWriter(OUT_PATH, engine="openpyxl") as writer:
    table2_fmt.to_excel(writer, sheet_name="Table2_eval")
    confusion_df.to_excel(writer, sheet_name="Confusion_eval")
    table3_pair.to_excel(writer, sheet_name="Table3_pair_level")

    # Basic formatting
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    wb = writer.book
    for ws_name in wb.sheetnames:
        ws = wb[ws_name]

        # header
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # index column
        for cell in ws["A"]:
            if cell.row == 1:
                continue
            cell.font = Font(bold=True)
            cell.alignment = Alignment(vertical="center", wrap_text=True)

        # column widths
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            for row in range(1, ws.max_row + 1):
                v = ws.cell(row=row, column=col_idx).value
                if v is None:
                    continue
                max_len = max(max_len, len(str(v)))
            ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 45)

        ws.freeze_panes = "B2" if ws.max_column > 1 else "A2"

print(f"Saved: {OUT_PATH.resolve()}")
print("Prediction columns used:", pred_cols)
print("Manual non-missing rows for Table 2:", int(df[manual_col].notna().sum()))
print("Full rows for Table 3:", len(df))

